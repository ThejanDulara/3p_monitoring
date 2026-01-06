import io
from datetime import datetime, time
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd

from extractor import extract_schedule_grid
from monitoring import find_unmatched_records
from storage import put_extract, get_extract, put_result, get_result

app = Flask(__name__)
CORS(app)

# ---------- helpers ----------

def make_json_safe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert datetime / date / time objects to strings
    so Flask jsonify() does not crash.
    """
    df = df.copy()

    for col in df.columns:
        # datetime64 columns
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.strftime("%Y-%m-%d %H:%M:%S")

        # object columns that may contain time objects
        else:
            df[col] = df[col].apply(
                lambda x: x.strftime("%H:%M:%S") if isinstance(x, time)
                else x.strftime("%Y-%m-%d %H:%M:%S") if isinstance(x, datetime)
                else x
            )

    return df


def df_preview(df: pd.DataFrame, limit=200):
    df = make_json_safe(df)
    cols = list(df.columns)
    rows = df.head(limit).fillna("").to_dict(orient="records")
    return {
        "columns": cols,
        "rows": rows,
        "totalRows": int(len(df))
    }


def to_excel_bytes_from_df(df: pd.DataFrame):
    out = io.BytesIO()
    export_df = df.copy()

    # prefer Date_dt as Excel date if exists
    if "Date_dt" in export_df.columns:
        export_df["Date"] = export_df["Date_dt"]

    export_df = export_df.drop(columns=["Date_dt"], errors="ignore")

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Extracted Row Data")
        ws = writer.book["Extracted Row Data"]

        if "Date" in export_df.columns:
            date_col_idx = export_df.columns.get_loc("Date") + 1
            for r in range(2, len(export_df) + 2):
                cell = ws.cell(row=r, column=date_col_idx)
                if cell.value:
                    cell.number_format = "DD/MM/YYYY"

    out.seek(0)
    return out


# ---------- routes ----------

@app.get("/api/health")
def health():
    return {"ok": True}


@app.post("/api/schedule/sheets")
def schedule_sheets():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "file is required"}), 400

    import openpyxl
    tmp = io.BytesIO(f.read())
    wb = openpyxl.load_workbook(tmp, data_only=True)
    sheets = [n for n in wb.sheetnames if n != "Final KPIs"]

    return jsonify({"sheets": sheets})


@app.post("/api/extract")
def extract():
    f = request.files.get("file")
    sheet = request.form.get("sheet", "")
    channel = request.form.get("channel", "")
    advertiser = request.form.get("advertiser", "")

    if not f or not sheet:
        return jsonify({"error": "file and sheet are required"}), 400

    import tempfile
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmpf:
        tmpf.write(f.read())
        tmp_path = tmpf.name

    df = extract_schedule_grid(tmp_path, sheet, channel, advertiser)
    token = put_extract(df, meta={
        "sheet": sheet,
        "channel": channel,
        "advertiser": advertiser
    })

    return jsonify({
        "token": token,
        "preview": df_preview(df)
    })


@app.get("/api/extract/download/<token>")
def download_extracted(token):
    item = get_extract(token)
    if not item:
        return jsonify({"error": "invalid or expired token"}), 404

    excel_bytes = to_excel_bytes_from_df(item["df"])

    return send_file(
        excel_bytes,
        as_attachment=True,
        download_name="extracted_schedule.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.post("/api/monitor")
def monitor():
    token = request.form.get("token", "")
    ro_number = request.form.get("ro_number", "")
    f = request.files.get("nilson")

    if not token or not ro_number or not f:
        return jsonify({"error": "token, ro_number, nilson file are required"}), 400

    item = get_extract(token)
    if not item:
        return jsonify({"error": "invalid or expired token"}), 404

    schedule_df = item["df"].copy()
    nilson_df = pd.read_excel(f)

    unmatched_df, updated_nilson_df = find_unmatched_records(
        schedule_df, nilson_df, ro_number
    )

    matched_count = int(
        (updated_nilson_df.get("RO Number") == ro_number).sum()
    ) if "RO Number" in updated_nilson_df.columns else 0

    summary = {
        "totalScheduleSpots": int(len(schedule_df)),
        "totalUnmatched": int(len(unmatched_df)),
        "totalMatchedInNilson": matched_count
    }

    job_id = put_result(unmatched_df, updated_nilson_df, summary=summary)

    return jsonify({
        "job_id": job_id,
        "summary": summary,
        "unmatchedPreview": df_preview(unmatched_df),
        "nilsonPreview": df_preview(updated_nilson_df)
    })


@app.get("/api/monitor/download/<job_id>/<which>")
def download_monitor_files(job_id, which):
    item = get_result(job_id)
    if not item:
        return jsonify({"error": "invalid or expired job"}), 404

    if which == "unmatched":
        df = item["unmatched"]
        name = "unmatched_data.csv"
    elif which == "nilson":
        df = item["nilson"]
        name = "nilson.csv"
    else:
        return jsonify({"error": "which must be unmatched or nilson"}), 400

    buf = io.BytesIO()
    df.to_csv(buf, index=False, encoding="utf-8-sig")
    buf.seek(0)

    return send_file(buf, as_attachment=True, download_name=name, mimetype="text/csv")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
