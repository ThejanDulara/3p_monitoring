from datetime import datetime
import openpyxl
import pandas as pd

COL_MAP = {
    "program": "program",
    "com_name": "com name",
    "duration": "duration",
    "language": "language",
    "time": "time",
    "rate_card": "nrate",
    "negotiated": "ncost"
}

OUTPUT_HEADERS = [
    "Program", "Time", "Language", "Dur", "Rate Card Rate",
    "Negotiated Rate", "Date", "Commercial Name", "Channel", "Advertiser"
]


def get_merged_value(sheet, row, col):
    cell = sheet.cell(row=row, column=col)
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return cell.value


def list_valid_sheets(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    names = [n for n in wb.sheetnames if n != "Final KPIs"]
    return names


def extract_schedule_grid(xlsx_path: str, sheet_name: str, channel: str, advertiser: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = wb[sheet_name]

    # 1) find header row containing "Program" in col A (rows 1..24)
    header_row = None
    indices = {}

    for r in range(1, 25):
        v = sheet.cell(row=r, column=1).value
        if v is None:
            continue
        if "program" in str(v).lower():
            header_row = r
            break

    if not header_row:
        raise ValueError("Could not find program header row.")

    # 2) map fixed columns A-R (1..18)
    for c in range(1, 19):
        cell_val = sheet.cell(row=header_row, column=c).value
        cell_val = "" if cell_val is None else str(cell_val).lower()
        for key, target in COL_MAP.items():
            if target in cell_val:
                indices[key] = c

    # 3) map date columns starting from S (19)
    month_row = header_row - 3
    date_num_row = header_row - 1
    date_cols = {}

    for c in range(19, sheet.max_column + 1):
        d_num = sheet.cell(row=date_num_row, column=c).value
        if d_num is None:
            break
        m_year = get_merged_value(sheet, month_row, c)  # e.g. "Jan - 2026"
        date_cols[c] = f"{d_num} {m_year}"  # e.g. "21 Jan - 2026"

    rows = []

    # 4) extract one entry per spot
    for r in range(header_row + 1, sheet.max_row + 1):
        prog = sheet.cell(row=r, column=indices.get("program", 1)).value
        if not prog:
            continue
        prog_s = str(prog).lower()
        if any(x in prog_s for x in ["total", "benefit", "bonus", "commercial"]):
            continue

        for col_idx, date_str in date_cols.items():
            spot_count = sheet.cell(row=r, column=col_idx).value
            if spot_count and isinstance(spot_count, (int, float)) and spot_count > 0:
                for _ in range(int(spot_count)):
                    rows.append({
                        "Program": prog,
                        "Commercial Name": sheet.cell(row=r, column=indices.get("com_name", 2)).value,
                        "Dur": sheet.cell(row=r, column=indices.get("duration", 3)).value,
                        "Language": sheet.cell(row=r, column=indices.get("language", 4)).value,
                        "Time": sheet.cell(row=r, column=indices.get("time", 6)).value,
                        "Date": date_str,
                        "Rate Card Rate": sheet.cell(row=r, column=indices.get("rate_card", 8)).value,
                        "Negotiated Rate": sheet.cell(row=r, column=indices.get("negotiated", 9)).value,
                        "Channel": channel,
                        "Advertiser": advertiser
                    })

    df = pd.DataFrame(rows)

    # Normalize Date: parse "21 Jan - 2026" -> datetime, but keep as datetime for Excel formatting
    def parse_date(raw):
        try:
            return datetime.strptime(str(raw).strip(), "%d %b - %Y")
        except Exception:
            return pd.NaT

    if not df.empty:
        df["Date_dt"] = df["Date"].apply(parse_date)
    else:
        df["Date_dt"] = pd.Series(dtype="datetime64[ns]")

    # Keep display column "Date" in DD/MM/YYYY if possible (for table)
    df["Date"] = df["Date_dt"].dt.strftime("%d/%m/%Y")

    # order columns like your export
    df = df[[
        "Program", "Time", "Language", "Dur",
        "Rate Card Rate", "Negotiated Rate", "Date",
        "Commercial Name", "Channel", "Advertiser", "Date_dt"
    ]]

    return df
